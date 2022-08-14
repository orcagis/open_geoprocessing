#-------------------------------------------------------------------------------
# Name:        gis_output.py
# Purpose:     Gets coordinates for point layers
#
# Author:      Justin Hawley (justin@orcagis.com)
#
# Created:     08/04/2022
#
# Interpreter: /fastgisapi/venv/bin/python3
#-------------------------------------------------------------------------------

import csv
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter

# data structure cannot be changed once sent to this function, must be list of dicts
def export_csv(json, out_name, out_path, headers = True):
    out_name = '{}.csv'.format(out_name)
    if out_path is not None:
        out_name = os.path.join(out_path, out_name)
    # get columns, only grabs whats in first row
    columns = []
    for line in json:
        for key, val in line.items():
            columns.append(key)
        break
    with open(out_name, 'w', newline='') as csv_file:
            writer = csv.DictWriter(csv_file, fieldnames=columns)
            if headers:
                writer.writeheader()
            for key in json:
                writer.writerow(key)

def export_xlsx(json, out_name, out_path, headers = True):
    sheet_name = out_name
    out_name = '{}.xlsx'.format(out_name)
    if out_path is not None:
        out_name = os.path.join(out_path, out_name)
    columns = []
    for line in json:
        for key, val in line.items():
            columns.append(key)
        break
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    if headers:
        ws.append(columns)
    for dict in json:
        values = (dict[k] for k in columns)
        ws.append(values)
    wb.save(out_name)


def main():
    pass
if __name__ == '__main__':
    main()